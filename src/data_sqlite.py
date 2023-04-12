import openpyxl
import sqlite3
import time
from sqlite3 import Error
import pandas as pd


class Data_Sqlite:
    def __init__(self):
        self.db_file = "F:\\freelancer\\Tool_Ethereum\\data\\result.db"
        self.xlsx_file = "F:\\freelancer\\Tool_Ethereum\\data\\result.xlsx"

        global conn
        while True:
            try:
                conn = sqlite3.connect(self.db_file)
                break
            except Error as e:
                print(e)
                time.sleep(0.2)
                continue

        self.conn = conn

    def create_table(self, create_table_sql):
        try:
            c = self.conn.cursor()
            c.execute(create_table_sql)
        except Error as e:
            print(e)

    def create_wallet(self, wallet):
        sql = ''' INSERT INTO wallets(account,balance,status)
                  VALUES(?,?,?) '''
        cur = self.conn.cursor()
        cur.execute(sql, wallet)
        self.conn.commit()
        return cur.lastrowid

    def sqlite_to_xlsx(self, sheet_name='Sheet1'):
        global sql
        try:
            # Làm sạch file excel
            df = pd.read_excel(self.xlsx_file)
            df.drop(df.index, inplace=True)
            df.to_excel(self.xlsx_file, index=False)

            sql = 'SELECT  id, account, balance, status FROM wallets'

            cur = self.conn.cursor()
            cur.execute(sql)
            data = cur.fetchall()

            wb = openpyxl.load_workbook(self.xlsx_file)
            sheet = wb[sheet_name]

            sheet.cell(row=1, column=3, value="ID")
            sheet.cell(row=1, column=4, value="ACCOUNT")
            sheet.cell(row=1, column=5, value="BALANCE")
            sheet.cell(row=1, column=6, value="STATUS")

            # print(data[1])
            # sheet.cell(row=3, column=6, value=data[3 - 2][6 - 3])
            for i in range(2, len(data) + 2):
                for j in range(3, 7):
                    sheet.cell(row=i, column=j, value=data[i-2][j-3])

            # Căn giữa toàn bộ
            ws = wb.active
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center")

            wb.close()
            wb.save(self.xlsx_file)
        except Exception as e:
            print(e)


# def main():   
#     #
#     sql_create_wallets_table = """CREATE TABLE IF NOT EXISTS wallets (
#                                     id integer PRIMARY KEY,
#                                     account text NOT NULL,
#                                     balance integer,
#                                     status text NOT NULL
#                                 );"""
#
#     # create a database connection
#     Data_Sqlite().create_table(sql_create_wallets_table)
#     # print("Done")
#     # project_id = data_sqlite(database).create_config(("2", '2015-01-01'))
#     #
#     # data_sqlite(database).create_ldplayer(("LD1", 0, "Kwin68", project_id, "15-12-2022"))
#     # data_sqlite(1.25).create_ldplayer(("LD70-2", 123456, "Block", 70, "15-12-2022"))
#     # data_sqlite(1.25).create_ldplayer(("LD70-3", 123456, "Kww68", 70, "15-12-2022"))
#
#     # print("done")
#     #
#     # # print(list(data_sqlite(1, "Kwin68").count_record())[0])
#     # # data_sqlite(1.25).sqlite_to_xlsx()
#     # print(len(data_sqlite(1, "Kwin68").select_all_ldplayers()))
#     # Data_Sqlite().sqlite_to_xlsx()
#     # print("Done")
# #
# #
# main()